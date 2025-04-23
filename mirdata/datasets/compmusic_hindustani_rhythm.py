"""CompMusic Hindustani Rhythm Dataset Loader

.. admonition:: Dataset Info
    :class: dropdown

    CompMusic Hindustani Rhythm Dataset is a rhythm annotated test corpus for automatic rhythm analysis tasks in Hindustani Music.
    The collection consists of audio excerpts from the CompMusic Hindustani research corpus, manually annotated time aligned markers
    indicating the progression through the taal cycle, and the associated taal related metadata. A brief description of the dataset
    is provided below.

    For a brief overview and audio examples of taals in Hindustani music, please see: http://compmusic.upf.edu/examples-taal-hindustani

    The dataset contains the following data:

    **AUDIO:** The pieces are chosen from the CompMusic Hindustani music collection. The pieces were chosen in four popular taals of Hindustani music,
    which encompasses a majority of Hindustani khyal music. The pieces were chosen include a mix of vocal and instrumental recordings, new and old
    recordings, and to span three lays. For each taal, there are pieces in dhrut (fast), madhya (medium) and vilambit (slow) lays (tempo class). All
    pieces have Tabla as the percussion accompaniment. The excerpts are two minutes long. Each piece is uniquely identified using the MBID of the recording.
    The pieces are stereo, 160 kbps, mp3 files sampled at 44.1 kHz. The audio is also available as wav files for experiments.

    **SAM, VIBHAAG AND THE MAATRAS:** The primary annotations are audio synchronized time-stamps indicating the different metrical positions in the taal cycle.
    The sam and matras of the cycle are annotated. The annotations were created using Sonic Visualizer by tapping to music and manually correcting the taps.
    Each annotation has a time-stamp and an associated numeric label that indicates the position of the beat marker in the taala cycle. The annotations and the
    associated metadata have been verified for correctness and completeness by a professional Hindustani musician and musicologist. The long thick lines show
    vibhaag boundaries. The numerals indicate the matra number in cycle. In each case, the sam (the start of the cycle, analogous to the downbeat) are indicated
    using the numeral 1.

    **METADATA:** For each excerpt, the taal and the lay of the piece are recorded. Each excerpt can be uniquely identified and located with the MBID of the
    recording, and the relative start and end times of the excerpt within the whole recording. A separate 5 digit taal based unique ID is also provided for each
    excerpt as a double check. The artist, release, the lead instrument, and the raag of the piece are additional editorial metadata obtained from the release.
    There are optional comments on audio quality and annotation specifics.

    The dataset consists of excerpts with a wide tempo range from 10 MPM (matras per minute) to 370 MPM. To study any effects of the tempo class, the full dataset
    (HMDf) is also divided into two other subsets - the long cycle subset (HMDl) consisting of vilambit (slow) pieces with a median tempo between 10-60 MPM, and the
    short cycle subset (HMDs) with madhyalay (medium, 60-150 MPM) and the drut lay (fast, 150+ MPM).

    **Possible uses of the dataset:** Possible tasks where the dataset can be used include taal, sama and beat tracking, tempo estimation and tracking, taal recognition,
    rhythm based segmentation of musical audio, audio to score/lyrics alignment, and rhythmic pattern discovery.

    **Dataset organization:** The dataset consists of audio, annotations, an accompanying spreadsheet providing additional metadata, a MAT-file that has identical
    information as the spreadsheet, and a dataset description document.

    The annotations files of this dataset are shared with the following license: Creative Commons Attribution Non Commercial Share Alike 4.0 International

"""

import os
import csv
import logging
import librosa
import numpy as np

from mirdata import annotations, core, io
from smart_open import open


try:
    from openpyxl import load_workbook as get_xlxs
except ImportError:
    logging.error(
        "In order to use CompMusic Hindustani Music Rhythm you must have openpyxl installed. "
        "Please reinstall mirdata using `pip install 'mirdata[compmusic_hindustani_rhythm]'"
    )
    raise

BIBTEX = """
@inproceedings{Srinivasamurthy2016,
    author = {Srinivasamurthy, Ajay and Holzapfel, Andre and Cemgil, Ali and Serra, Xavier},
    year = {2016},
    month = {03},
    pages = {76-80},
    title = {A generalized Bayesian model for tracking long metrical cycles in acoustic music signals},
    doi = {10.1109/ICASSP.2016.7471640}
}
"""

INDEXES = {
    "default": "1.0",
    "test": "sample",
    "1.0": core.Index(
        filename="compmusic_hindustani_rhythm_full_index_1.0.json",
        url="https://zenodo.org/records/14007893/files/compmusic_hindustani_rhythm_full_index_1.0.json?download=1",
        checksum="1b66dfd109bf453626be0b7352c9fa3a",
    ),
    "sample": core.Index(
        filename="compmusic_hindustani_rhythm_full_index_1.0_sample.json"
    ),
}

REMOTES = None

LICENSE_INFO = (
    "Creative Commons Attribution Non Commercial Share Alike 4.0 International."
)

DOWNLOAD_INFO = """The files of this dataset are shared under request. Please go to: https://zenodo.org/record/1264742 and request access, stating
    the research-related use you will give to the dataset. Once the access is granted (it may take, at most, one day or two), please download 
    the dataset with the provided Zenodo link and uncompress and store the datasets to a desired location, and use such location to initialize the 
    dataset as follows: compmusic_hindustani_rhythm = mirdata.initialize("compmusic_hindustani_rhythm", data_home="/path/to/home/folder/of/dataset").
    """


class Track(core.Track):
    """CompMusic Hindustani Music Rhythm class

    Args:
        track_id (str): track id of the track
        data_home (str): Local path where the dataset is stored. default=None
            If `None`, looks for the data in the default directory, `~/mir_datasets`

    Attributes:
        audio_path (str): path to audio file
        beats_path (srt): path to beats file
        meter_path (srt): path to meter file

    Cached Properties:
        beats (BeatData): beats annotation
        meter (string): meter annotation
        mbid (string): MusicBrainz ID
        name (string): name of the recording in the dataset
        artist (string): artists name
        release (string): release name
        lead_instrument_code (string): code for the load instrument
        taala (string): taala annotation
        raaga (string): raaga annotation
        laya (string): laya annotation
        num_of_beats (int): number of beats in annotation
        num_of_samas (int): number of samas in annotation
        median_matra_period (float): median matra per period
        median_matras_per_min (float): median matras per minute
        median_ISI (float): median ISI
        median_avarts_per_min (float): median avarts per minute
    """

    def __init__(
        self,
        track_id,
        data_home,
        dataset_name,
        index,
        metadata,
    ):
        super().__init__(
            track_id,
            data_home,
            dataset_name,
            index,
            metadata,
        )

        # Audio path
        self.audio_path = self.get_path("audio")

        # Annotations paths
        self.beats_path = self.get_path("beats")
        self.meter_path = self.get_path("meter")

    @core.cached_property
    def beats(self):
        return load_beats(self.beats_path)

    @core.cached_property
    def meter(self):
        return load_meter(self.meter_path)

    @core.cached_property
    def mbid(self):
        return self._track_metadata.get("mbid")

    @core.cached_property
    def name(self):
        return self._track_metadata.get("name")

    @core.cached_property
    def artist(self):
        return self._track_metadata.get("artist")

    @core.cached_property
    def release(self):
        return self._track_metadata.get("release")

    @core.cached_property
    def lead_instrument_code(self):
        return self._track_metadata.get("lead_instrument_code")

    @core.cached_property
    def taala(self):
        return self._track_metadata.get("taala")

    @core.cached_property
    def raaga(self):
        return self._track_metadata.get("raaga")

    @core.cached_property
    def laya(self):
        return self._track_metadata.get("laya")

    @core.cached_property
    def num_of_beats(self):
        return self._track_metadata.get("num_of_beats")

    @core.cached_property
    def num_of_samas(self):
        return self._track_metadata.get("num_of_samas")

    @core.cached_property
    def median_matra_period(self):
        return self._track_metadata.get("median_matra_period")

    @core.cached_property
    def median_matras_per_min(self):
        return self._track_metadata.get("median_matras_per_min")

    @core.cached_property
    def median_ISI(self):
        return self._track_metadata.get("median_ISI")

    @core.cached_property
    def median_avarts_per_min(self):
        return self._track_metadata.get("median_avarts_per_min")

    @property
    def audio(self):
        """The track's audio

        Returns:
           * np.ndarray - audio signal
           * float - sample rate

        """
        return load_audio(self.audio_path)


# no decorator here because of https://github.com/librosa/librosa/issues/1267
def load_audio(audio_path):
    """Load an audio file.

    Args:
        audio_path (str): path to audio file

    Returns:
        * np.ndarray - the mono audio signal
        * float - The sample rate of the audio file

    """
    if audio_path is None:
        return None
    return librosa.load(audio_path, sr=44100, mono=False)


@io.coerce_to_string_io
def load_beats(fhandle):
    """Load beats

    Args:
        fhandle (str or file-like): Local path where the beats annotation is stored.

    Returns:
        BeatData: beat annotations

    """
    beat_times = []
    beat_positions = []

    reader = csv.reader(fhandle, delimiter=",")
    for line in reader:
        beat_times.append(float(line[0]))
        beat_positions.append(int(line[1]))

    if not beat_times or beat_times[0] == -1.0:
        return None

    return annotations.BeatData(
        np.array(beat_times), "s", np.array(beat_positions), "bar_index"
    )


@io.coerce_to_string_io
def load_meter(fhandle):
    """Load meter

    Args:
        fhandle (str or file-like): Local path where the meter annotation is stored.

    Returns:
        float: meter annotation

    """
    reader = csv.reader(fhandle, delimiter=",")
    return next(reader)[0]


@core.docstring_inherit(core.Dataset)
class Dataset(core.Dataset):
    """
    The compmusic_hindustani_rhythm dataset

    """

    def __init__(self, data_home=None, version="default"):
        super().__init__(
            data_home,
            version,
            name="compmusic_hindustani_rhythm",
            track_class=Track,
            bibtex=BIBTEX,
            indexes=INDEXES,
            remotes=REMOTES,
            license_info=LICENSE_INFO,
            download_info=DOWNLOAD_INFO,
        )

    @core.cached_property
    def _metadata(self):
        metadata_path = os.path.join(self.data_home, "HMR_1.0", "HMDf.xlsx")

        metadata = {}
        try:
            with open(metadata_path, "rb") as fhandle:
                reader = get_xlxs(fhandle)
                reade = reader["HMDf"]
                rows = 0

                # Get actual number of rows
                for _, row in enumerate(reade, 1):
                    if not all(col.value is None for col in row):
                        rows += 1

                # Get actual columns
                columns = []
                for cell in reade[1]:
                    if cell.value:
                        columns.append(cell.value)

                for row in range(2, rows + 1):
                    metadata[str(reade.cell(row, 1).value)] = {
                        "mbid": reade.cell(row, 3).value,
                        "name": reade.cell(row, 4).value,
                        "artist": reade.cell(row, 5).value,
                        "release": reade.cell(row, 6).value,
                        "lead_instrument_code": reade.cell(row, 7).value,
                        "raaga": reade.cell(row, 8).value,
                        "taala": reade.cell(row, 9).value,
                        "laya": reade.cell(row, 10).value,
                        "num_of_beats": int(reade.cell(row, 13).value),
                        "num_of_samas": int(reade.cell(row, 14).value),
                        "median_matra_period": float(reade.cell(row, 16).value),
                        "median_matras_per_min": round(
                            60 / float(reade.cell(row, 16).value), 2
                        ),
                        "median_ISI": float(reade.cell(row, 16).value) * 16,
                        "median_avarts_per_min": round(
                            60 / (float(reade.cell(row, 16).value) * 16), 2
                        ),
                    }

        except FileNotFoundError:
            raise FileNotFoundError("Metadata not found. Did you run .download()?")

        return metadata
