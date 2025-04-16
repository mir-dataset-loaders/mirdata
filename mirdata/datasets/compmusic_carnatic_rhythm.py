"""CompMusic Carnatic Rhythm Dataset Loader

.. admonition:: Dataset Info
    :class: dropdown

    CompMusic Carnatic Rhythm Dataset is a rhythm annotated test corpus for automatic rhythm analysis tasks in Carnatic Music.
    The collection consists of audio excerpts from the CompMusic Carnatic research corpus, manually annotated time aligned markers
    indicating the progression through the taala cycle, and the associated taala related metadata. A brief description of the dataset
    is provided below. For a brief overview and audio examples of taalas in Carnatic music, please see:
    http://compmusic.upf.edu/examples-taala-carnatic

    The dataset contains the following data:

    **AUDIO:** The pieces are chosen from the CompMusic Carnatic music collection. The pieces were chosen in four popular taalas of
    Carnatic music, which encompasses a majority of Carnatic music. The pieces were chosen include a mix of vocal and instrumental recordings,
    new and old recordings, and to span a wide variety of forms. All pieces have a percussion accompaniment, predominantly Mridangam. The
    excerpts are full length pieces or a part of the full length pieces. There are also several different pieces by the same artist (or release
    group), and multiple instances of the same composition rendered by different artists. Each piece is uniquely identified using the MBID of the
    recording. The pieces are stereo, 160 kbps, mp3 files sampled at 44.1 kHz.

    **SAMA AND BEATS:** The primary annotations are audio synchronized time-stamps indicating the different metrical positions in the taala cycle.
    The annotations were created using Sonic Visualizer by tapping to music and manually correcting the taps. Each annotation has a time-stamp and
    an associated numeric label that indicates the position of the beat marker in the taala cycle. The marked positions in the taala cycle are shown
    with numbers, along with the corresponding label used. In each case, the sama (the start of the cycle, analogous to the downbeat) are indicated
    using the numeral 1.

    **METADATA:** For each excerpt, the taala of the piece, edupu (offset of the start of the piece, relative to the sama, measured in aksharas)
    of the composition, and the kalai (the cycle length scaling factor) are recorded. Each excerpt can be uniquely identified and located with the
    MBID of the recording, and the relative start and end times of the excerpt within the whole recording. A separate 5 digit taala based unique ID
    is also provided for each excerpt as a double check. The artist, release, the lead instrument, and the raaga of the piece are additional
    editorial metadata obtained from the release. A flag indicates if the excerpt is a full piece or only a part of a full piece. There are optional
    comments on audio quality and annotation specifics.

    Possible uses of the dataset: Possible tasks where the dataset can be used include taala, sama and beat tracking, tempo estimation and tracking,
    taala recognition, rhythm based segmentation of musical audio, structural segmentation, audio to score/lyrics alignment, and rhythmic pattern
    discovery.

    Dataset organization: The dataset consists of audio, annotations, an accompanying spreadsheet providing additional metadata. For a detailed
    description of the organization, please see the README in the dataset.

    Data Subset: A subset of this dataset consisting of 118 two minute excerpts of music is also available. The content in the subset is equaivalent
    and is separately distributed for a quicker testing of algorithms and approaches.

    The annotations files of this dataset are shared with the following license:
    Creative Commons Attribution Non Commercial Share Alike 4.0 International

"""

import os
import csv
import logging
import librosa
import numpy as np

from mirdata import annotations, core, io
from smart_open import open


try:
    from openpyxl import load_workbook as get_xlxs
except ImportError:
    logging.error(
        "In order to use CompMusic Carnatic Music Rhythm you must have openpyxl installed. "
        "Please reinstall mirdata using `pip install 'mirdata[compmusic_carnatic_rhythm]'"
    )
    raise

BIBTEX = """
@article{srinivasamurthy_2014,
  title={Particle Filters for Efficient Meter Tracking with Dynamic Bayesian Networks},
  author={Srinivasamurthy, A. and Holzapfel, A. and Cemgil, A. T. and Serra, X.},
  journal={In Proceedings of the 16th International Society for Music Information Retrieval Conference (ISMIR)},
  pages={197--203}
  year={2015}
}
"""

INDEXES = {
    "default": "full_dataset_1.0",
    "full_dataset": "full_dataset_1.0",
    "subset": "subset_1.0",
    "test": "sample",
    "full_dataset_1.0": core.Index(
        filename="compmusic_carnatic_rhythm_full_index_1.0.json",
        url="https://zenodo.org/records/14007971/files/compmusic_carnatic_rhythm_full_index_1.0.json?download=1",
        checksum="22d13adb87a3e9f3b5162cb2f73b638f",
    ),
    "subset_1.0": core.Index(
        filename="compmusic_carnatic_rhythm_subset_index_1.0.json",
        url="https://zenodo.org/records/14007996/files/compmusic_carnatic_rhythm_subset_index_1.0.json?download=1",
        checksum="05e8e5570d0f57fb36d75a50538e2afb",
    ),
    "sample": core.Index(
        filename="compmusic_carnatic_rhythm_subset_index_1.0_sample.json"
    ),
}

REMOTES = None

LICENSE_INFO = (
    "Creative Commons Attribution Non Commercial Share Alike 4.0 International."
)

DOWNLOAD_INFO = """The files of this dataset are shared under request. Please go to: https://zenodo.org/record/1264394 and request access, stating
    the research-related use you will give to the dataset. Once the access is granted (it may take, at most, one day or two), please download 
    the dataset with the provided Zenodo link and uncompress the two zip files: CMR_full_dataset_1.0.zip and CMR_subset_1.0.zip. You don't need 
    to re-arrange or change the folder structure of these two versions, the dataloader is designed to work with the provided file organization. 
    Therefore, simply uncompress and store the datasets to a desired location, and use such location to initialize the dataset as follows: 
    
    compmusic_carnatic_rhythm = mirdata.initialize("compmusic_carnatic_rhythm", data_home="/path/to/home/folder/of/dataset").
    """


class Track(core.Track):
    """CompMusic Carnatic Music Rhythm class

    Args:
        track_id (str): track id of the track
        data_home (str): Local path where the dataset is stored. default=None
            If `None`, looks for the data in the default directory, `~/mir_datasets`

    Attributes:
        audio_path (str): path to audio file
        beats_path (srt): path to beats file
        meter_path (srt): path to meter file

    Cached Properties:
        beats (BeatData): beats annotation
        meter (string): meter annotation
        mbid (string): MusicBrainz ID
        name (string): name of the recording in the dataset
        artist (string): artists name
        release (string): release name
        lead_instrument_code (string): code for the load instrument
        taala (string): taala annotation
        raaga (string): raaga annotation
        num_of_beats (int): number of beats in annotation
        num_of_samas (int): number of samas in annotation

    """

    def __init__(
        self,
        track_id,
        data_home,
        dataset_name,
        index,
        metadata,
    ):
        super().__init__(
            track_id,
            data_home,
            dataset_name,
            index,
            metadata,
        )

        # Audio path
        self.audio_path = self.get_path("audio")

        # Annotations paths
        self.beats_path = self.get_path("beats")
        self.meter_path = self.get_path("meter")

    @core.cached_property
    def beats(self):
        return load_beats(self.beats_path)

    @core.cached_property
    def meter(self):
        return load_meter(self.meter_path)

    @core.cached_property
    def mbid(self):
        return self._track_metadata.get("mbid")

    @core.cached_property
    def name(self):
        return self._track_metadata.get("name")

    @core.cached_property
    def artist(self):
        return self._track_metadata.get("artist")

    @core.cached_property
    def release(self):
        return self._track_metadata.get("release")

    @core.cached_property
    def lead_instrument_code(self):
        return self._track_metadata.get("lead_instrument_code")

    @core.cached_property
    def taala(self):
        return self._track_metadata.get("taala")

    @core.cached_property
    def raaga(self):
        return self._track_metadata.get("raaga")

    @core.cached_property
    def num_of_beats(self):
        return self._track_metadata.get("num_of_beats")

    @core.cached_property
    def num_of_samas(self):
        return self._track_metadata.get("num_of_samas")

    @property
    def audio(self):
        """The track's audio

        Returns:
           * np.ndarray - audio signal
           * float - sample rate

        """
        return load_audio(self.audio_path)


# no decorator here because of https://github.com/librosa/librosa/issues/1267
def load_audio(audio_path):
    """Load an audio file.

    Args:
        audio_path (str): path to audio file

    Returns:
        * np.ndarray - the mono audio signal
        * float - The sample rate of the audio file

    """
    if audio_path is None:
        return None
    return librosa.load(audio_path, sr=44100, mono=False)


@io.coerce_to_string_io
def load_beats(fhandle):
    """Load beats

    Args:
        fhandle (str or file-like): Local path where the beats annotation is stored.

    Returns:
        BeatData: beat annotations

    """
    beat_times = []
    beat_positions = []

    reader = csv.reader(fhandle, delimiter=",")
    for line in reader:
        beat_times.append(float(line[0]))
        beat_positions.append(int(line[1]))

    if not beat_times or beat_times[0] == -1.0:
        return None

    return annotations.BeatData(
        np.array(beat_times), "s", np.array(beat_positions), "bar_index"
    )


@io.coerce_to_string_io
def load_meter(fhandle):
    """Load meter

    Args:
        fhandle (str or file-like): Local path where the meter annotation is stored.

    Returns:
        float: meter annotation

    """
    reader = csv.reader(fhandle, delimiter=",")
    return next(reader)[0]


@core.docstring_inherit(core.Dataset)
class Dataset(core.Dataset):
    """
    The compmusic_carnatic_rhythm dataset

    """

    def __init__(self, data_home=None, version="default"):
        super().__init__(
            data_home,
            version,
            name="compmusic_carnatic_rhythm",
            track_class=Track,
            bibtex=BIBTEX,
            indexes=INDEXES,
            remotes=REMOTES,
            license_info=LICENSE_INFO,
            download_info=DOWNLOAD_INFO,
        )

    @core.cached_property
    def _metadata(self):
        if self.version == "full_dataset_1.0":
            metadata_path = os.path.join(
                self.data_home, "CMR_full_dataset_1.0", "CMRfullDataset.xlsx"
            )

        else:
            metadata_path = os.path.join(
                self.data_home, "CMR_subset_1.0", "CMRdataset.xlsx"
            )

        metadata = {}
        try:
            with open(metadata_path, "rb") as fhandle:
                reader = get_xlxs(fhandle)
                if self.version == "full_dataset_1.0":
                    reade = reader["Carnatic"]
                    rows = 0

                    # Get actual number of rows
                    for _, row in enumerate(reade, 1):
                        if not all(col.value is None for col in row):
                            rows += 1

                    # Get actual columns
                    columns = []
                    for cell in reade[1]:
                        if cell.value:
                            columns.append(cell.value)

                    for row in range(2, rows + 1):
                        metadata[str(reade.cell(row, 1).value)] = {
                            "mbid": reade.cell(row, 2).value,
                            "name": reade.cell(row, 3).value,
                            "artist": reade.cell(row, 4).value,
                            "release": reade.cell(row, 5).value,
                            "lead_instrument_code": reade.cell(row, 6).value,
                            "taala": reade.cell(row, 7).value,
                            "raaga": reade.cell(row, 8).value,
                            "num_of_beats": int(reade.cell(row, 13).value),
                            "num_of_samas": int(reade.cell(row, 14).value),
                        }

                else:
                    reader = reader.active
                    rows = 0

                    # Get actual number of rows
                    for _, row in enumerate(reader, 1):
                        if not all(col.value is None for col in row):
                            rows += 1

                    # Get actual columns
                    columns = []
                    for cell in reader[1]:
                        if cell.value:
                            columns.append(cell.value)

                    rows_it = range(2, rows) if rows > 2 else range(2, rows + 1)
                    for row in rows_it:
                        metadata[str(reader.cell(row, 2).value)] = {
                            "mbid": reader.cell(row, 3).value,
                            "name": reader.cell(row, 4).value,
                            "artist": reader.cell(row, 5).value,
                            "release": reader.cell(row, 6).value,
                            "lead_instrument_code": reader.cell(row, 7).value,
                            "taala": reader.cell(row, 8).value,
                            "raaga": reader.cell(row, 9).value,
                            "num_of_beats": int(reader.cell(row, 10).value),
                            "num_of_samas": int(reader.cell(row, 11).value),
                        }

        except FileNotFoundError:
            raise FileNotFoundError("Metadata not found. Did you run .download()?")

        return metadata
